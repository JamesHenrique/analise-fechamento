"""
Análise de Ociosidade - Relive
KPIs para decisão: abrir 8h, fechar 19h, eliminar sábados
Gera: PPTX, PDF, HTML
"""
import openpyxl
from collections import Counter, defaultdict
from datetime import datetime, date
import json

# ============================================================
# 1. CARREGAR DADOS
# ============================================================
BASE = 'G:/Outros computadores/NOTEBOOK1/James/projeto_relive/Relatório/Relatório de abertura-fechamento'

# AGENDA SÁBADO
wb1 = openpyxl.load_workbook(f'{BASE}/AGENDA SÁBADO 01.2026 a 30.06.2026 (1).xlsx')
ws1 = wb1.active
agenda_sabado = []
for row in ws1.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    data_str = row[0]
    hora_str = row[2]
    profissional = str(row[3]).strip() if row[3] else ''
    paciente = str(row[4]).strip() if row[4] else ''
    tipo = str(row[5]).strip() if row[5] else ''
    if isinstance(data_str, str):
        try:
            data = datetime.strptime(data_str, '%d/%m/%Y')
        except:
            try:
                data = datetime.strptime(data_str, '%Y-%m-%d')
            except:
                continue
    elif isinstance(data_str, datetime):
        data = data_str
    else:
        continue
    agenda_sabado.append({
        'data': data,
        'hora': str(hora_str).strip() if hora_str else '',
        'profissional': profissional,
        'paciente': paciente,
        'tipo': tipo,
        'origem': 'sabado'
    })

# ATENDIMENTOS antes 8h / após 19h
wb2 = openpyxl.load_workbook(f'{BASE}/atendimentos 01.2026 a 30.06.2027 antes das 8h e término após 19h (1).xlsx')
ws2 = wb2.active
atendimentos_fora = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[1] is None:
        continue
    data_str = row[1]
    hora_str = row[2]
    profissional = str(row[3]).strip() if row[3] else ''
    paciente = str(row[4]).strip() if row[4] else ''
    tipo = str(row[5]).strip() if row[5] else ''
    if isinstance(data_str, str):
        try:
            data = datetime.strptime(data_str, '%d/%m/%Y')
        except:
            try:
                data = datetime.strptime(data_str, '%Y-%m-%d')
            except:
                continue
    elif isinstance(data_str, datetime):
        data = data_str
    else:
        continue

    # Classificar: antes 8h ou após 19h
    hora_val = str(hora_str).strip()
    categoria = 'antes_8h' if hora_val < '08:00' else 'apos_19h'

    atendimentos_fora.append({
        'data': data,
        'hora': hora_val,
        'profissional': profissional,
        'paciente': paciente,
        'tipo': tipo,
        'categoria': categoria,
        'origem': categoria
    })

# ============================================================
# 2. ANÁLISE - KPIs
# ============================================================

# 2.1 Sábados
total_sabados = len(agenda_sabado)
sabados_por_mes = Counter()
sabados_por_profissional = Counter()
sabados_por_tipo = Counter()
sabados_por_hora = Counter()
sabados_pacientes_unicos = set()

for a in agenda_sabado:
    mes = a['data'].strftime('%Y-%m')
    sabados_por_mes[mes] += 1
    sabados_por_profissional[a['profissional']] += 1
    sabados_por_tipo[a['tipo']] += 1
    sabados_por_hora[a['hora']] += 1
    sabados_pacientes_unicos.add(a['paciente'])

# Contar sábados únicos com atendimento
sabados_unicos = set()
for a in agenda_sabado:
    sabados_unicos.add(a['data'].strftime('%Y-%m-%d'))

# Total de sábados no período (01/01/2026 a 30/06/2026)
# Contar sábados reais
from datetime import timedelta
d_inicio = date(2026, 1, 1)
d_fim = date(2026, 6, 30)
total_sabados_periodo = 0
d = d_inicio
while d <= d_fim:
    if d.weekday() == 5:  # 5 = sábado
        total_sabados_periodo += 1
    d += timedelta(days=1)

# 2.2 Antes 8h
atendimentos_antes_8h = [a for a in atendimentos_fora if a['categoria'] == 'antes_8h']
total_antes_8h = len(atendimentos_antes_8h)
antes_8h_por_mes = Counter()
antes_8h_por_profissional = Counter()
antes_8h_por_tipo = Counter()
antes_8h_pacientes_unicos = set()

for a in atendimentos_antes_8h:
    mes = a['data'].strftime('%Y-%m')
    antes_8h_por_mes[mes] += 1
    antes_8h_por_profissional[a['profissional']] += 1
    antes_8h_por_tipo[a['tipo']] += 1
    antes_8h_pacientes_unicos.add(a['paciente'])

# 2.3 Após 19h
atendimentos_apos_19h = [a for a in atendimentos_fora if a['categoria'] == 'apos_19h']
total_apos_19h = len(atendimentos_apos_19h)
apos_19h_por_mes = Counter()
apos_19h_por_profissional = Counter()
apos_19h_por_tipo = Counter()
apos_19h_pacientes_unicos = set()

for a in atendimentos_apos_19h:
    mes = a['data'].strftime('%Y-%m')
    apos_19h_por_mes[mes] += 1
    apos_19h_por_profissional[a['profissional']] += 1
    apos_19h_por_tipo[a['tipo']] += 1
    apos_19h_pacientes_unicos.add(a['paciente'])

# 2.4 Análise combinada
total_geral = total_sabados + total_antes_8h + total_apos_19h

# Dias úteis no período para estimar ocupação 7h-8h
dias_uteis = 0
d = d_inicio
while d <= d_fim:
    if d.weekday() < 5:  # seg-sex
        dias_uteis += 1
    d += timedelta(days=1)

# Média atendimentos por dia antes 8h
media_antes_8h_dia = total_antes_8h / dias_uteis if dias_uteis > 0 else 0

# Média atendimentos por sábado
media_sabado_dia = total_sabados / len(sabados_unicos) if sabados_unicos else 0

# Top 5 profissionais mais afetados (todos cenários)
todos_atendimentos = agenda_sabado + atendimentos_fora
prof_impacto = Counter()
for a in todos_atendimentos:
    prof_impacto[a['profissional']] += 1

# ============================================================
# 3. COMPILAR RESULTADOS
# ============================================================

kpis = {
    'total_sabados': total_sabados,
    'total_antes_8h': total_antes_8h,
    'total_apos_19h': total_apos_19h,
    'total_geral': total_geral,
    'sabados_unicos': len(sabados_unicos),
    'total_sabados_periodo': total_sabados_periodo,
    'sabados_sem_atendimento': total_sabados_periodo - len(sabados_unicos),
    'taxa_ocupacao_sabados': len(sabados_unicos) / total_sabados_periodo * 100,
    'media_antes_8h_dia': media_antes_8h_dia,
    'media_sabado_dia': media_sabado_dia,
    'dias_uteis': dias_uteis,
    'sabados_por_mes': dict(sabados_por_mes.most_common()),
    'antes_8h_por_mes': dict(antes_8h_por_mes.most_common()),
    'apos_19h_por_mes': dict(apos_19h_por_mes.most_common()),
    'sabados_por_profissional': dict(sabados_por_profissional.most_common(10)),
    'antes_8h_por_profissional': dict(antes_8h_por_profissional.most_common(10)),
    'apos_19h_por_profissional': dict(apos_19h_por_profissional.most_common(10)),
    'sabados_por_tipo': dict(sabados_por_tipo.most_common(10)),
    'antes_8h_por_tipo': dict(antes_8h_por_tipo.most_common(10)),
    'apos_19h_por_tipo': dict(apos_19h_por_tipo.most_common(10)),
    'sabados_pacientes_unicos': len(sabados_pacientes_unicos),
    'antes_8h_pacientes_unicos': len(antes_8h_pacientes_unicos),
    'apos_19h_pacientes_unicos': len(apos_19h_pacientes_unicos),
    'top10_profissionais_impactados': dict(prof_impacto.most_common(10)),
    'sabados_por_hora': dict(sorted(sabados_por_hora.items())),
}

print("=== KPIs CALCULADOS ===")
print(json.dumps(kpis, indent=2, ensure_ascii=False, default=str))

# Salvar KPIs para usar nos scripts de geração
with open(f'{BASE}/kpis.json', 'w', encoding='utf-8') as f:
    json.dump(kpis, f, ensure_ascii=False, default=str, indent=2)

print("\nKPIs salvos em kpis.json")
print(f"Total sábados no período: {total_sabados_periodo}")
print(f"Sábados com atendimento: {len(sabados_unicos)}")
print(f"Sábados SEM atendimento: {total_sabados_periodo - len(sabados_unicos)}")
print(f"Taxa ocupação sábados: {kpis['taxa_ocupacao_sabados']:.1f}%")
print(f"Total atendimentos antes 8h: {total_antes_8h}")
print(f"Média atendimentos/dia antes 8h: {media_antes_8h_dia:.2f}")
print(f"Total atendimentos após 19h: {total_apos_19h}")
print(f"Total atendimentos sábados: {total_sabados}")
print(f"Média atendimentos/sábado: {media_sabado_dia:.2f}")
